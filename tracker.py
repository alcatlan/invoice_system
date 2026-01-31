import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import os

def registrar_en_excel(fecha, factura_no, total):
    archivo_excel = "registro_contable.xlsx"
    
    try:
        # 📂 1. Obtener o crear el libro
        if not os.path.exists(archivo_excel):
            wb = Workbook()
            hoja = wb.active
            hoja.title = "Historial_Facturas"
            # Encabezados en negrita
            hoja.append(["Factura No", "Fecha", "Total (CAD)"])
            for cell in hoja[1]:
                cell.font = Font(bold=True)
        else:
            wb = openpyxl.load_workbook(archivo_excel)
            hoja = wb.active

        # 📝 2. Añadir los datos
        hoja.append([factura_no, fecha, total])
        
        # 💰 3. Aplicar formato de moneda a la celda del total (Columna C)
        ultima_fila = hoja.max_row
        celda_monto = hoja.cell(row=ultima_fila, column=3)
        celda_monto.number_format = '"$"#,##0.00'
        
        # 💾 4. Guardar cambios
        wb.save(archivo_excel)
        print(f"📊 Registro exitoso en Excel: Factura {factura_no}")

        # 📈 5. Mostrar resumen de ganancias acumuladas
        ganancia_total = 0.0
        for fila in range(2, hoja.max_row + 1):
            valor = hoja.cell(row=fila, column=3).value
            try:
                ganancia_total += float(valor)
            except (ValueError, TypeError):
                continue
        
        print(f"💰 Ganancia total acumulada hasta hoy: ${ganancia_total:,.2f} CAD")

    except PermissionError:
        print(f"\n⚠️ ERROR: No se pudo acceder a '{archivo_excel}'.")
        print("❌ Asegúrate de CERRAR el archivo de Excel antes de ejecutar el programa.")